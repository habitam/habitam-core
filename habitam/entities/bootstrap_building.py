# -*- coding: utf-8 -*-
'''
This file is part of Habitam.

Habitam is free software: you can redistribute it and/or modify
it under the terms of the GNU Affero General Public License as 
published by the Free Software Foundation, either version 3 of 
the License, or (at your option) any later version.

Habitam is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU Affero General Public License for more details.

You should have received a copy of the GNU Affero General Public
License along with Habitam. If not, see 
<http://www.gnu.org/licenses/>.
    
Created on Aug 12, 2013

@author: Stefan Guna
'''
from datetime import date, datetime
from habitam.entities.models import ApartmentGroup, Person, Apartment, \
    AccountLink, CollectingFund
from habitam.financial.models import Account
from openpyxl.datavalidation import DataValidation, ValidationType
from openpyxl.reader.excel import load_workbook
from openpyxl.style import Fill, Color
from openpyxl.workbook import Workbook
import tempfile


def __create_building(name, daily_penalty=None, close_day=None,
                    payment_due_days=None):
    default_account = Account.objects.create(type='std')
    penalties_account = Account.objects.create(type='penalties')
    building = ApartmentGroup.objects.create(name=name, type='building',
                            default_account=default_account,
                            penalties_account=penalties_account,
                            daily_penalty=daily_penalty,
                            close_day=close_day,
                            payment_due_days=payment_due_days)
    
    AccountLink.objects.create(holder=building, account=default_account)
    default_account.name = building.__unicode__()
    default_account.save()
    
    AccountLink.objects.create(holder=building, account=penalties_account)
    penalties_account.name = building.__unicode__()
    penalties_account.save()
    
    return building

def __create_staircase(name, parent, daily_penalty=None, close_day=None,
                    payment_due_days=None):
    staircase = ApartmentGroup.objects.create(name=name, parent=parent,
                            type='stair', daily_penalty=daily_penalty,
                            close_day=close_day,
                            payment_due_days=payment_due_days)
    staircase.save()
    
    return staircase


def bootstrap_building(user_license, name, staircases, apartments,
                       apartment_offset, daily_penalty, close_day,
                       payment_due_days):
    per_staircase = apartments / staircases
    remainder = apartments % staircases
    
    building = __create_building(name=name,
                    daily_penalty=daily_penalty,
                    close_day=close_day,
                    payment_due_days=payment_due_days)
    ap_idx = apartment_offset
    today = date.today()
    for i in range(staircases):
        staircase = __create_staircase(parent=building, name=str(i + 1))
        for j in range(per_staircase):
            name = str(ap_idx)
            owner = Person.bootstrap_owner(name)
            Apartment.objects.create(name=name, parent=staircase,
                                     owner=owner, no_penalties_since=today)
            ap_idx = ap_idx + 1
        if i + 1 < staircases:
            continue
        for j in range(remainder):
            name = str(ap_idx)
            owner = Person.bootstrap_owner(name)
            Apartment.objects.create(name=name, parent=staircase,
                                     owner=owner, no_penalties_since=today)
            ap_idx = ap_idx + 1
    if user_license != None:
        user_license.add_entity(building, ApartmentGroup)
    else:
        building.save()
        
    f = CollectingFund.objects.create(billed=building, quota_type='equally',
                        name='Fond reparații', money_type='cash',
                        account_type='repairs')
    f.account.type = 'repairs'
    f.account.save()
    
    f = CollectingFund.objects.create(billed=building, quota_type='noquota',
                        name='Fond rulment', money_type='cash',
                        account_type='rulment')
    f.account.type = 'rulment'
    f.account.save()

    return building


def initial_operations(building, sums, dates):
    assert sums
    assert dates
    assert (k in dates.keys() for k in sums.keys())
    assert len(sums) == len(dates)
    
    name = u'Sold inițial'
    f = CollectingFund.objects.create(archived=True,
                        archive_date=datetime.fromtimestamp(0),
                        billed=building, quota_type='noquota',
                        name=name)
    
    for ap, s in sums.iteritems():
        no = 'Sold inițial ' + str(ap)
        f.new_inbound_operation(amount=s * -1, no=no, ap_sums={ap:s * -1},
                        date=dates[ap])
        
def load_initial_operations(building, xlsx):
    wb = load_workbook(xlsx)
    ws = wb.get_sheet_by_name(u'Solduri inițiale')
    
    ap_ids = map(lambda ap: ap.id, building.apartments())
    num_rows = ws.get_highest_row()
    if num_rows - 1 != len(ap_ids):
        raise Exception(u'Fișier invalid')
    sums = {}
    dates = {}
    
    for row in range(1, num_rows):
        ap_id = int(ws.cell(row=row, column=0).value)
        if not ap_id in ap_ids:
            raise Exception(u'Fișier invalid')
        s = ws.cell(row=row, column=2).value
        d = ws.cell(row=row, column=3).value
        if s == None or d == None:
            continue
        sums[ap_id] = s
        dates[ap_id] = d
    
    if not sums:
        raise Exception(u'Nu ați introdus nici o valoare')
    
    initial_operations(building, sums, dates)
        
def initial_operations_template(building):
    wb = Workbook()
    ws = wb.worksheets[0]
    ws.title = u'Solduri inițiale'
    ws.cell(row=0, column=0).value = 'ID intern'
    ws.cell(row=0, column=1).value = 'Apartament'
    ws.cell(row=0, column=2).value = 'Sold'
    ws.cell(row=0, column=3).value = 'Data'
    
    
    def attr_validator(attr):
        l = (str(getattr(ap, attr)) for ap in building.apartments())
        formula = '"' + ','.join(l) + '"'
        validator = DataValidation(ValidationType.LIST, formula1=formula,
                                    allow_blank=False)
        return validator
    
    id_validator = attr_validator('id')
    ws.add_data_validation(id_validator)
    name_validator = attr_validator('name')
    ws.add_data_validation(name_validator)
    
    decimal_validator = DataValidation(ValidationType.DECIMAL, allow_blank=True)
    decimal_validator.set_error_message('Introduceți un număr zecimal', 'Eroare de validare')
    ws.add_data_validation(decimal_validator)
    
    date_validator = DataValidation(ValidationType.DATE, allow_blank=True)
    date_validator.set_error_message('Introduceți o dată', 'Eroare de validare')
    ws.add_data_validation(date_validator)

    row = 1
    now = datetime.now()
    for ap in building.apartments():
        cell_id = ws.cell(row=row, column=0)
        cell_id.value = ap.id
        id_validator.add_cell(cell_id)
        
        cell_name = ws.cell(row=row, column=1)
        cell_name.value = ap.name
        name_validator.add_cell(cell_name)
        
        cell_balance = ws.cell(row=row, column=2)
        cell_balance.style.fill.fill_type = Fill.FILL_SOLID
        cell_balance.style.fill.start_color.index = Color.YELLOW
        decimal_validator.add_cell(cell_balance)
        
        cell_date = ws.cell(row=row, column=3)
        cell_date.value = now
        cell_date.style.number_format.format_code = 'yyyy-mm-dd'
        cell_date.style.fill.fill_type = Fill.FILL_SOLID
        cell_date.style.fill.start_color.index = Color.YELLOW
        date_validator.add_cell(cell_date)
        
        row += 1
    
    ws.column_dimensions['A'].visible = False
    ws.column_dimensions['B'].width = 20.0
    ws.column_dimensions['C'].width = 20.0
    ws.column_dimensions['D'].width = 20.0
    temp = tempfile.NamedTemporaryFile()
    wb.save(temp)
    return temp
    
